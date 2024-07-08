# from uiform import Ui_MainWindow
# from PyQt6.QtWidgets import QApplication,QMainWindow,QPushButton
import ip
import os
import openpyxl
from PyQt6.QtWidgets import *
from uiform import Ui_MainWindow
from login import Ui_MainLogin

#___________________________________________________________________________________________________
#Đăng nhập
class login_Window(QMainWindow,Ui_MainLogin):
    def __init__(self):
        super(login_Window,self).__init__()
        ip.uic.loadUi('login.ui',self)
        self.btn_login.clicked.connect(self.loginApp)
        
        
    def loginApp(self):
        tk = self.username.text()
        mk = self.password.text()
        kt = ip.ConnectDB.loginApp(tk, mk)
        if kt:
            widget.setFixedHeight(750)
            widget.setFixedWidth(850)
            widget.move(250, 100)
            widget.setCurrentIndex(1)
        else:
            ip.QMessageBox.information(self, "Thông báo", "Sai tài khoản hoặc mật khẩu!")
    

#___________________________________________________________________________________________________
#Chức năng chính
class MySideBar(QMainWindow, Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.setWindowTitle("SideBar Menu")
        self.icon_name_widget.setHidden(True)

# ___________________________________________________________________________________________________
#btn dang xuat
      
# ___________________________________________________________________________________________________
#Btn chuyển trang

        self.sanPham_1.clicked.connect(self.switch_to_sanPham_page)
        self.sanpham_2.clicked.connect(self.switch_to_sanPham_page)

        self.nhanvien_1.clicked.connect(self.switch_to_nhanVien_page)
        self.nhanvien_2.clicked.connect(self.switch_to_nhanVien_page)

        self.donhang_1.clicked.connect(self.switch_to_donhang_page)
        self.donhang_2.clicked.connect(self.switch_to_donhang_page)

        self.doanhthu_1.clicked.connect(self.switch_to_doanhThu_page)
        self.doanhthu_2.clicked.connect(self.switch_to_doanhThu_page)

        self.dangXuat_1.clicked.connect(self.switch_to_DangXuatPage)
        self.dangXuat_2.clicked.connect(self.switch_to_DangXuatPage)

        
# ___________________________________________________________________________________________________
# #Btn sản phẩm
        self.showSanPham()
        self.tbl_sanPham.cellClicked.connect(self.tbl_sanPham_clicked)
        self.btn_themsp.clicked.connect(self.themSanPham)
        self.btn_xoasp.clicked.connect(self.deleteSanPham)
        self.btn_suasp.clicked.connect(self.updateSanPham)
        self.btn_timKiemsp.clicked.connect(self.showSanPham_TimKiem)   

# ___________________________________________________________________________________________________
#Btn nhân viên
        self.showNhanVien()
        self.tbl_nhanVien.cellClicked.connect(self.tbl_nhanvien_clicked)
        self.btn_themnv.clicked.connect(self.themNhanvien)
        self.btn_suanv.clicked.connect(self.updateNhanvien)
        self.btn_xoanv.clicked.connect(self.deleteNhanvien)
        self.btn_timKiemnv.clicked.connect(self.showNhanvien_TimKiem)

# ___________________________________________________________________________________________________
# #Btn đơn hàng 
        self.showDonHang()
        self.tbl_donHang.cellClicked.connect(self.tbl_donhang_clicked)
        self.btn_themdh.clicked.connect(self.themDonHang)
        self.btn_suadh.clicked.connect(self.updateDonHang)
        self.btn_xoadh.clicked.connect(self.deleteDonHang)
        self.btn_timKiemdh.clicked.connect(self.showDonHang_TimKiem)
# ___________________________________________________________________________________________________
#btn Doanh thu
        self.showDoanhThu()
        self.thongKe.clicked.connect(self.tinhTong)
        self.xuatExcel.clicked.connect(self.xuatFileExcel)
# ___________________________________________________________________________________________________

    def switch_to_DangXuatPage(self):
        reply = ip.QMessageBox.question(self, 'Xác nhận', 
                                     "Bạn có muốn đăng xuất không?", 
                                     ip.QMessageBox.StandardButton.Yes | ip.QMessageBox.StandardButton.No, 
                                     ip.QMessageBox.StandardButton.No)

        if reply == ip.QMessageBox.StandardButton.Yes:
            print("dang xuat")
            
            widget.setFixedHeight(550)
            widget.setFixedWidth(790)
            widget.move(250, 100)
            widget.setCurrentIndex(0)
            print("Đăng xuất ngay!")
        else:
            print("Hủy đăng xuất") 

#Hàm chuyển trang
    def switch_to_sanPham_page(self):
        self.stackedWidget.setCurrentIndex(0)

    def switch_to_nhanVien_page(self):
        self.stackedWidget.setCurrentIndex(1)

    def switch_to_donhang_page(self):
        self.showDonHang()
        self.stackedWidget.setCurrentIndex(3)

    def switch_to_doanhThu_page(self):
        self.stackedWidget.setCurrentIndex(4)
    
    def click_DangXuat(self):
        self.close()

# ___________________________________________________________________________________________________
#Hàm quản lý sản phẩm
    def showSanPham(self):
        self.tbl_sanPham.setRowCount(ip.DAL_SanPham.showSanPham().__len__())
        self.tbl_sanPham.setColumnCount(6)
        self.tbl_sanPham.setHorizontalHeaderLabels(['Mã sản phẩm', 'Tên sản phẩm', 'Số lượng', 'Giá nhập', 'Giá bán', 'Mô tả'])
        self.tbl_sanPham.setColumnWidth(1, 150)
        self.tbl_sanPham.setColumnWidth(5, 150)
        table_row = 0
        for row in ip.DAL_SanPham.showSanPham():
            self.tbl_sanPham.setItem(table_row, 0, ip.QTableWidgetItem(str(row[0])))
            self.tbl_sanPham.setItem(table_row, 1, ip.QTableWidgetItem(str(row[1])))
            self.tbl_sanPham.setItem(table_row, 2, ip.QTableWidgetItem(str(row[2])))
            self.tbl_sanPham.setItem(table_row, 3, ip.QTableWidgetItem(str(row[3])))
            self.tbl_sanPham.setItem(table_row, 4, ip.QTableWidgetItem(str(row[4])))
            self.tbl_sanPham.setItem(table_row, 5, ip.QTableWidgetItem(str(row[5])))
            table_row += 1

    def tbl_sanPham_clicked(self,row,column):
        self.masp.setText(self.tbl_sanPham.item(row,0).text())
        self.tensp.setText(self.tbl_sanPham.item(row,1).text())
        self.soLuong.setText(self.tbl_sanPham.item(row,3).text())
        self.giaNhap.setText(self.tbl_sanPham.item(row,2).text())
        self.giaBan.setText(self.tbl_sanPham.item(row,4).text())
        self.moTa.setText(self.tbl_sanPham.item(row,5).text())

    def themSanPham(self):
        masp = self.masp.text()
        tensp = self.tensp.text()
        soLuong= self.soLuong.text()
        giaNhap = self.giaNhap.text()
        giaBan = self.giaBan.text()
        moTa = self.moTa.text()
        kt = ip.DAL_SanPham.addSanPham(masp, tensp, soLuong, giaNhap, giaBan, moTa)
        if kt == 1:
            ip.QMessageBox.information(self, "Thông báo", "Thêm thành công!")
            self.showSanPham()
        else:
            ip.QMessageBox.information(self, "Thông báo", "Thêm không thành công!")

    def updateSanPham(self):
        masp = self.masp.text()
        tensp = self.tensp.text()
        soLuong = self.soLuong.text()
        giaNhap = self.giaNhap.text()
        giaban = self.giaBan.text()
        moTa = self.moTa.text()
        kt = ip.DAL_SanPham.updateSanPham(masp, tensp, soLuong,giaNhap,giaban, moTa)
        if kt == 1:
            ip.QMessageBox.information(self, "Thông báo", "Cập nhật thành công!")
            self.showSanPham()
        else:
            ip.QMessageBox.information(self, "Thông báo", "Cập nhật không thành công!")
    
    def showSanPham_TimKiem(self):
        tensp = self.timKiemsp.text()
        self.tbl_sanPham.setRowCount(ip.DAL_SanPham.timKiem(tensp).__len__())
        self.tbl_sanPham.setColumnCount(6)
        self.tbl_sanPham.setHorizontalHeaderLabels(["Mã sản phẩm", "Tên sản phẩm", "Số lượng", "Giá nhập", "Giá bán", "Mô tả"])
        self.tbl_sanPham.setColumnWidth(1, 150)
        self.tbl_sanPham.setColumnWidth(5, 150)
        table_row = 0
        for row in ip.DAL_SanPham.timKiem(tensp):
            self.tbl_sanPham.setItem(table_row, 0, ip.QTableWidgetItem(str(row[0])))
            self.tbl_sanPham.setItem(table_row, 1, ip.QTableWidgetItem(str(row[1])))
            self.tbl_sanPham.setItem(table_row, 2, ip.QTableWidgetItem(str(row[2])))
            self.tbl_sanPham.setItem(table_row, 3, ip.QTableWidgetItem(str(row[3])))
            self.tbl_sanPham.setItem(table_row, 4, ip.QTableWidgetItem(str(row[4])))
            self.tbl_sanPham.setItem(table_row, 5, ip.QTableWidgetItem(str(row[5])))
            table_row += 1
    def deleteSanPham(self):
        maGV = self.masp.text()
        kt = ip.DAL_SanPham.deleteSanPham(maGV)
        if kt == 1:
            ip.QMessageBox.information(self, "Thông báo", "Xóa thành công!")
            self.showSanPham()
        else:
            ip.QMessageBox.information(self, "Thông báo", "Xóa không thành công!")

# ___________________________________________________________________________________________________   
#Hàm qly nhân viên
    def showNhanVien(self):
        self.tbl_nhanVien.setRowCount(ip.DAL_NhanVien.showNhanVien().__len__())
        self.tbl_nhanVien.setColumnCount(7)
        self.tbl_nhanVien.setHorizontalHeaderLabels(['Mã nhân viên', 'Tên nhân viên', 'Username', 'Password', 'Giới tính', 'Số điện thoại', 'Chức vụ'])
        self.tbl_nhanVien.setColumnWidth(1, 150)
        self.tbl_nhanVien.setColumnWidth(6, 150)
        table_row = 0
        for row in ip.DAL_NhanVien.showNhanVien():
            self.tbl_nhanVien.setItem(table_row, 0, ip.QTableWidgetItem(str(row[0])))
            self.tbl_nhanVien.setItem(table_row, 1, ip.QTableWidgetItem(str(row[1])))
            self.tbl_nhanVien.setItem(table_row, 2, ip.QTableWidgetItem(str(row[2])))
            self.tbl_nhanVien.setItem(table_row, 3, ip.QTableWidgetItem(str(row[3])))
            self.tbl_nhanVien.setItem(table_row, 4, ip.QTableWidgetItem(str(row[4])))
            self.tbl_nhanVien.setItem(table_row, 5, ip.QTableWidgetItem(str(row[5])))
            self.tbl_nhanVien.setItem(table_row, 6, ip.QTableWidgetItem(str(row[6])))
            table_row += 1
    def tbl_nhanvien_clicked(self,row,column):
        self.manv.setText(self.tbl_nhanVien.item(row,0).text())
        self.tennv.setText(self.tbl_nhanVien.item(row,1).text())
        self.username.setText(self.tbl_nhanVien.item(row,2).text())
        self.password.setText(self.tbl_nhanVien.item(row,3).text())
        self.gender.setCurrentText(self.tbl_nhanVien.item(row, 4).text())
        self.sdt_nhanVien.setText(self.tbl_nhanVien.item(row,5).text())
        self.chucVu.setCurrentText(self.tbl_nhanVien.item(row,6).text())

    def themNhanvien(self):
        manv = self.manv.text()
        tennv = self.tennv.text()
        username = self.username.text()
        password = self.password.text()
        gender = self.gender.currentText()
        sdt = self.sdt_nhanVien.text()
        chucVu = self.chucVu.currentText()
        kt = ip.DAL_NhanVien.addNhanVien(manv, tennv, username, password, gender, sdt, chucVu)
        if kt == 1:
            ip.QMessageBox.information(self, "Thông báo", "Thêm thành công!")
            self.showNhanVien()
        else:
            ip.QMessageBox.information(self, "Thông báo", "Thêm không thành công!")

    def updateNhanvien(self):
        manv = self.manv.text()
        tennv = self.tennv.text()
        username = self.username.text()
        password = self.password.text()
        gender = self.gender.currentText()
        sdt = self.sdt_nhanVien.text()
        chucVu = self.chucVu.currentText()
        kt = ip.DAL_NhanVien.updateNhanVien(manv, tennv, username, password, gender, sdt, chucVu)
        if kt == 1:
            ip.QMessageBox.information(self, "Thông báo", "Cập nhật thành công!")
            self.showNhanVien()
        else:
            ip.QMessageBox.information(self, "Thông báo", "Cập nhật không thành công!")

    def deleteNhanvien(self):
        manv = self.manv.text()
        kt = ip.DAL_NhanVien.deleteNhanVien(manv)
        if kt == 1:
            ip.QMessageBox.information(self, "Thông báo", "Xóa thành công!")
            self.showNhanVien()
        else:
            ip.QMessageBox.information(self, "Thông báo", "Xóa không thành công!")

    def showNhanvien_TimKiem(self):
        tennv = self.timKiemnv.text()
        self.tbl_nhanVien.setRowCount(ip.DAL_NhanVien.timKiem(tennv).__len__())
        self.tbl_nhanVien.setColumnCount(7)
        self.tbl_nhanVien.setHorizontalHeaderLabels(["Mã sản phẩm", "Tên sản phẩm", "Số lượng", "Giá nhập", "Giá bán", "Mô tả"])
        self.tbl_nhanVien.setColumnWidth(1, 150)
        self.tbl_nhanVien.setColumnWidth(6, 150)
        table_row = 0
        for row in ip.DAL_NhanVien.timKiem(tennv):
            self.tbl_nhanVien.setItem(table_row, 0, ip.QTableWidgetItem(str(row[0])))
            self.tbl_nhanVien.setItem(table_row, 1, ip.QTableWidgetItem(str(row[1])))
            self.tbl_nhanVien.setItem(table_row, 2, ip.QTableWidgetItem(str(row[2])))
            self.tbl_nhanVien.setItem(table_row, 3, ip.QTableWidgetItem(str(row[3])))
            self.tbl_nhanVien.setItem(table_row, 4, ip.QTableWidgetItem(str(row[4])))
            self.tbl_nhanVien.setItem(table_row, 5, ip.QTableWidgetItem(str(row[5])))
            self.tbl_nhanVien.setItem(table_row, 6, ip.QTableWidgetItem(str(row[6])))
            table_row += 1

# ___________________________________________________________________________________________________
#Hàm qly đơn hàng
    def showDonHang(self):
        self.masp_dh.clear()
        for i in range(ip.DAL_SanPham.showSanPham().__len__()):
            self.masp_dh.addItem(str(ip.DAL_SanPham.showSanPham()[i][0]))
        self.tbl_donHang.setRowCount(ip.DAL_DonHang.showDonhang().__len__())
        self.tbl_donHang.setColumnCount(8)
        self.tbl_donHang.setHorizontalHeaderLabels(['Mã đơn hàng', 'Mã sản phẩm', 'Số lượng', 'Tên khách hàng', 'Địa chỉ', 'Số điện thoại', 'Tổng tiền','Trạng thái'])
        self.tbl_donHang.setColumnWidth(1, 150)
        self.tbl_donHang.setColumnWidth(7, 150)
        table_row = 0
        for row in ip.DAL_DonHang.showDonhang():
            self.tbl_donHang.setItem(table_row, 0, ip.QTableWidgetItem(str(row[0])))
            self.tbl_donHang.setItem(table_row, 1, ip.QTableWidgetItem(str(row[1])))
            self.tbl_donHang.setItem(table_row, 2, ip.QTableWidgetItem(str(row[2])))
            self.tbl_donHang.setItem(table_row, 3, ip.QTableWidgetItem(str(row[3])))
            self.tbl_donHang.setItem(table_row, 4, ip.QTableWidgetItem(str(row[4])))
            self.tbl_donHang.setItem(table_row, 5, ip.QTableWidgetItem(str(row[5])))
            self.tbl_donHang.setItem(table_row, 6, ip.QTableWidgetItem(str(row[6])))
            self.tbl_donHang.setItem(table_row, 7, ip.QTableWidgetItem(str(row[7])))
            table_row += 1

    def tbl_donhang_clicked(self,row,column):
        self.madh.setText(self.tbl_donHang.item(row,0).text())
        masp = self.tbl_donHang.item(row, 1).text()
        for i in range(self.masp_dh.count()):
            if(masp == self.masp_dh.itemText(i)):
                self.masp_dh.setCurrentIndex(i)
        self.soluong_dh.setText(self.tbl_donHang.item(row,2).text())
        self.ten_khach_hang_dh.setText(self.tbl_donHang.item(row,3).text())
        self.addressdh.setText(self.tbl_donHang.item(row,4).text())
        self.sdt_dh.setText(self.tbl_donHang.item(row,5).text())
        self.tongtien_dh.setText(self.tbl_donHang.item(row,6).text())
        self.trangthai_dh.setCurrentText(self.tbl_donHang.item(row,7).text())

    def themDonHang(self):
        madh = self.madh.text()
        masp = self.masp_dh.currentText()
        soLuong = int(self.soluong_dh.text())
        tenKh = self.ten_khach_hang_dh.text()
        adress = self.addressdh.text()
        sdt = self.sdt_dh.text()
        trangThai = self.trangthai_dh.currentText()
       

        giaBan = ip.DAL_SanPham.GiaSanPham(masp)[0][0]
        if giaBan is not None:
        # Tính tổng tiền
            tongTien = soLuong * giaBan
            self.tongtien_dh.setText(str(tongTien))  # Đặt giá trị vào ô tongTien
            self.tongtien_dh.setReadOnly(True)
           
        # Thêm đơn hàng
            kt = ip.DAL_DonHang.addDonHang(madh, masp, soLuong, tenKh, adress, sdt, tongTien, trangThai)
            if kt == 1:
                ip.QMessageBox.information(self, "Thông báo", "Thêm thành công!")
                self.showDonHang()
            else:
                ip.QMessageBox.information(self, "Thông báo", "Thêm không thành công!")
        else:
            ip.QMessageBox.information(self, "Thông báo", "Không tìm thấy giá bán cho mã sản phẩm đã chọn!")
        giaBan = 0
        self.showDonHang()
    def updateDonHang(self):
        madh = self.madh.text()
        masp = self.masp_dh.currentText()
        soLuong = int(self.soluong_dh.text())
        tenKh = self.ten_khach_hang_dh.text()
        adress = self.addressdh.text()
        sdt = self.sdt_dh.text()
        tongTien = self.tongtien_dh.text()
        trangThai = self.trangthai_dh.currentText()

        giaBan = ip.DAL_SanPham.GiaSanPham(masp)[0][0]
    
        if giaBan is not None:
        # Tính tổng tiền
            tongTien = soLuong * giaBan
            self.tongtien_dh.setText(str(tongTien))  # Đặt giá trị vào ô tongTien
        # Thêm đơn hàng
            kt = ip.DAL_DonHang.updateDonHang(madh, masp, soLuong, tenKh, adress, sdt, tongTien, trangThai)
            if kt == 1:
                ip.QMessageBox.information(self, "Thông báo", "Sửa thành công!")
                self.showDonHang()
            else:
                ip.QMessageBox.information(self, "Thông báo", "Sửa không thành công!")
        else:
            ip.QMessageBox.information(self, "Thông báo", "Không tìm thấy giá bán cho mã sản phẩm đã chọn!")
        giaBan = 0
    def deleteDonHang(self):
        madh = self.madh.text()
        kt = ip.DAL_DonHang.deleteDonHang(madh)
        if kt == 1:
            ip.QMessageBox.information(self, "Thông báo", "Xóa thành công!")
            self.showDonHang()
        else:
            ip.QMessageBox.information(self, "Thông báo", "Xóa không thành công!")     

    def showDonHang_TimKiem(self):
        tenkh = self.timKiemdh.text()
        self.tbl_donHang.setRowCount(ip.DAL_DonHang.timKiem(tenkh).__len__())
        self.tbl_donHang.setColumnCount(8)
        self.tbl_donHang.setHorizontalHeaderLabels(["Mã đơn hàng", "Mã sản phẩm", "Số lượng", "Tên khách hàng", "Địa chi", "Số điện thoại", "Tổng tiền", "Trang thái"])
        self.tbl_donHang.setColumnWidth(1, 150)
        self.tbl_donHang.setColumnWidth(7, 150)
        table_row = 0
        for row in ip.DAL_DonHang.timKiem(tenkh):
            self.tbl_donHang.setItem(table_row, 0, ip.QTableWidgetItem(str(row[0])))
            self.tbl_donHang.setItem(table_row, 1, ip.QTableWidgetItem(str(row[1])))
            self.tbl_donHang.setItem(table_row, 2, ip.QTableWidgetItem(str(row[2])))
            self.tbl_donHang.setItem(table_row, 3, ip.QTableWidgetItem(str(row[3])))
            self.tbl_donHang.setItem(table_row, 4, ip.QTableWidgetItem(str(row[4])))
            self.tbl_donHang.setItem(table_row, 5, ip.QTableWidgetItem(str(row[5])))
            self.tbl_donHang.setItem(table_row, 6, ip.QTableWidgetItem(str(row[6])))
            self.tbl_donHang.setItem(table_row, 7, ip.QTableWidgetItem(str(row[7])))
            table_row += 1
    
# _________________________________________________________________________________________________________________
#Doanh thu
   
    def showDoanhThu(self):
        self.tbl_tongdoanhthu.setRowCount(ip.DAL_DonHang.ThongKe().__len__())
        self.tbl_tongdoanhthu.setColumnCount(8)
        self.tbl_tongdoanhthu.setHorizontalHeaderLabels(['Mã đơn hàng', 'Mã sản phẩm', 'Số lượng', 'Tên khách hàng', 'Địa chỉ', 'Số điện thoại', 'Tổng tiền','Trạng thái'])
        self.tbl_tongdoanhthu.setColumnWidth(1, 150)
        self.tbl_tongdoanhthu.setColumnWidth(7, 150)
        table_row = 0
        TongTien = 0
        for row in ip.DAL_DonHang.ThongKe():
            self.tbl_tongdoanhthu.setItem(table_row, 0, ip.QTableWidgetItem(str(row[0])))
            self.tbl_tongdoanhthu.setItem(table_row, 1, ip.QTableWidgetItem(str(row[1])))
            self.tbl_tongdoanhthu.setItem(table_row, 2, ip.QTableWidgetItem(str(row[2])))
            self.tbl_tongdoanhthu.setItem(table_row, 3, ip.QTableWidgetItem(str(row[3])))
            self.tbl_tongdoanhthu.setItem(table_row, 4, ip.QTableWidgetItem(str(row[4])))
            self.tbl_tongdoanhthu.setItem(table_row, 5, ip.QTableWidgetItem(str(row[5])))
            self.tbl_tongdoanhthu.setItem(table_row, 6, ip.QTableWidgetItem(str(row[6])))
            self.tbl_tongdoanhthu.setItem(table_row, 7, ip.QTableWidgetItem(str(row[7])))
            table_row += 1
    
    def tinhTong(self):
        tinhTong =  ip.DAL_DonHang.tinhTongTien()
        self.tong_doanh_thu.setText(str(tinhTong))
        self.showDoanhThu()
        return tinhTong
    def xuatFileExcel(self):
        try:
            default_dir = os.path.expanduser("D:\Du an Python\Cuoi_ky")  
            file_path, _ = QFileDialog.getSaveFileName(self, 'Xuất Excel', default_dir, 'Excel Files (*.xlsx)')
            if not file_path:
                return
        
        # Tạo một workbook mới
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = 'Doanh thu'  # Đặt tên cho sheet
        
        # Đặt header cho file Excel
            headers = ['Mã đơn hàng', 'Mã sản phẩm', 'Số lượng', 'Tên khách hàng', 'Địa chỉ', 'Số điện thoại', 'Tổng tiền', 'Trạng thái']
            for col_index, header in enumerate(headers, start=1):
                sheet.cell(row=1, column=col_index, value=header)
        
        # Lấy dữ liệu từ bảng tbl_tongdoanhthu và điền vào file Excel
            for row_index, row_data in enumerate(ip.DAL_DonHang.ThongKe(), start=2):
                for col_index, cell_value in enumerate(row_data, start=1):
                    sheet.cell(row=row_index, column=col_index, value=str(cell_value))

            tinhTong = self.tinhTong()
            row_tinhTong = sheet.max_row + 2  # Tính vị trí hàng tiếp theo để ghi tổng doanh thu
            sheet.cell(row=row_tinhTong, column=1, value='Tổng doanh thu :')
            sheet.cell(row=row_tinhTong, column=2, value=str(tinhTong))
        # Lưu workbook vào file Excel
            wb.save(file_path)
        
        # Thông báo thành công cho người dùng
            QMessageBox.information(self, 'Thông báo', 'Xuất Excel thành công!')
    
        except Exception as e:
        # Xử lý các lỗi có thể xảy ra trong quá trình xuất Excel
            QMessageBox.critical(self, 'Lỗi', f'Có lỗi xảy ra: {str(e)}')

# _________________________________________________________________________________________________________________

app = ip.QApplication(ip.sys.argv)
widget = ip.QtWidgets.QStackedWidget()
login_f = login_Window()
mainGui_f = MySideBar()
widget.addWidget(login_f)
widget.addWidget(mainGui_f)
widget.setCurrentIndex(0)
widget.setFixedHeight(550)
widget.setFixedWidth(790)
widget.show()
app.exec()
